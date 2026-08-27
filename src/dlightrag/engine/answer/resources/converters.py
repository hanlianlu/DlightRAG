# Copyright 2025-2026 Hanlian Lu. SPDX-License-Identifier: Apache-2.0
"""Deterministic binary resource conversion with OOXML preflight.

Only HTML, CSV, PDF, DOCX, PPTX, and XLSX are admitted. MarkItDown runs with
plugins disabled and never fetches the network: DlightRAG hands it admitted bytes
and explicit :class:`StreamInfo`. A fresh converter is built per call so no
mutable state is shared between concurrent conversions. OOXML archives pass a
central-directory size preflight before any converter opens them, so a zip bomb
is rejected without decompressing attack-sized data. Embedded DOCX/PPTX images
are located through markdown-it image tokens and replaced with compact handles;
XLSX images are pulled from the workbook with their sheet/cell anchor.
"""

from __future__ import annotations

import asyncio
import base64
import binascii
import io
import zipfile
from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path

import openpyxl
from markdown_it import MarkdownIt
from markitdown import MarkItDown, StreamInfo
from openpyxl.utils import get_column_letter

from dlightrag.engine.answer.resources.models import ResourceRegistryError

# Physical archive-safety limits. These are internal decompression bounds, not
# the public attachment-size quotas, so an OOXML file that is admissible by byte
# size can still be rejected here if its internal expansion looks like a bomb.
_MAX_OOXML_ENTRIES = 10_000
_MAX_OOXML_ENTRY_BYTES = 100 * 1024 * 1024
_MAX_OOXML_TOTAL_BYTES = 512 * 1024 * 1024
_MAX_OOXML_EXPANSION_RATIO = 100

_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_PPTX_MIME = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ResourceConversionError(ResourceRegistryError):
    """Raised when an admitted resource cannot be converted deterministically."""


class UnsafeArchiveError(ResourceConversionError):
    """Raised when an OOXML archive exceeds a physical decompression limit."""


@dataclass(frozen=True)
class ExtractedVisual:
    """An image pulled out of a converted resource with its anchor and bytes."""

    handle_id: str
    anchor: str | None
    media_type: str
    data: bytes


@dataclass(frozen=True)
class ConvertedResource:
    """Deterministic text view plus the visuals extracted from one resource."""

    text: str
    visuals: tuple[ExtractedVisual, ...]


@dataclass(frozen=True)
class _Route:
    extension: str
    mimetype: str
    is_ooxml: bool
    is_xlsx: bool


_ROUTES: dict[str, _Route] = {
    ".html": _Route(".html", "text/html", False, False),
    ".htm": _Route(".htm", "text/html", False, False),
    ".csv": _Route(".csv", "text/csv", False, False),
    ".pdf": _Route(".pdf", "application/pdf", False, False),
    ".docx": _Route(".docx", _DOCX_MIME, True, False),
    ".pptx": _Route(".pptx", _PPTX_MIME, True, False),
    ".xlsx": _Route(".xlsx", _XLSX_MIME, True, True),
}
_MIME_TO_ROUTE: dict[str, _Route] = {route.mimetype: route for route in _ROUTES.values()}


def _resolve_route(filename: str | None, declared_mime: str | None) -> _Route | None:
    if filename:
        suffix = Path(filename).suffix.lower()
        route = _ROUTES.get(suffix)
        if route is not None:
            return route
    if declared_mime:
        return _MIME_TO_ROUTE.get(declared_mime.split(";", 1)[0].strip().lower())
    return None


def is_convertible(filename: str | None, declared_mime: str | None) -> bool:
    """Return whether an admitted suffix/MIME pair routes to a binary converter."""
    return _resolve_route(filename, declared_mime) is not None


async def convert_resource(
    content: bytes,
    *,
    filename: str | None,
    declared_mime: str | None,
) -> ConvertedResource:
    """Convert admitted binary *content* to text and extracted visuals off-loop."""
    route = _resolve_route(filename, declared_mime)
    if route is None:
        raise ResourceConversionError("resource type is not an admitted binary format")
    return await asyncio.to_thread(_convert_sync, content, route)


def _convert_sync(content: bytes, route: _Route) -> ConvertedResource:
    if route.is_ooxml:
        _preflight_ooxml(content)
    # One converter per call: registered converters and detector are never shared
    # across threads, so concurrent conversions cannot cross-contaminate.
    converter = MarkItDown(enable_plugins=False)
    try:
        result = converter.convert_stream(
            io.BytesIO(content),
            stream_info=StreamInfo(mimetype=route.mimetype, extension=route.extension),
            keep_data_uris=True,
        )
    except Exception as exc:  # noqa: BLE001 - surface any converter failure uniformly
        raise ResourceConversionError("resource conversion failed") from exc

    text, embedded = _extract_embedded_visuals(result.markdown)
    visuals = list(embedded)
    if route.is_xlsx:
        visuals.extend(_extract_xlsx_visuals(content))
    return ConvertedResource(text=text, visuals=tuple(visuals))


def _preflight_ooxml(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            sizes = [(info.file_size, info.compress_size) for info in archive.infolist()]
    except zipfile.BadZipFile as exc:
        raise ResourceConversionError("resource is not a valid OOXML archive") from exc
    _validate_archive_sizes(sizes)


def _validate_archive_sizes(sizes: list[tuple[int, int]]) -> None:
    """Reject zip-bomb shapes from central-directory ``(uncompressed, compressed)`` sizes."""
    if len(sizes) > _MAX_OOXML_ENTRIES:
        raise UnsafeArchiveError("OOXML archive has too many entries")
    total_uncompressed = 0
    total_compressed = 0
    for uncompressed, compressed in sizes:
        if uncompressed > _MAX_OOXML_ENTRY_BYTES:
            raise UnsafeArchiveError("OOXML entry exceeds the per-entry size limit")
        total_uncompressed += uncompressed
        total_compressed += compressed
    if total_uncompressed > _MAX_OOXML_TOTAL_BYTES:
        raise UnsafeArchiveError("OOXML total uncompressed size exceeds the limit")
    if total_compressed > 0 and total_uncompressed / total_compressed > _MAX_OOXML_EXPANSION_RATIO:
        raise UnsafeArchiveError("OOXML expansion ratio exceeds the limit")


def _extract_embedded_visuals(text: str) -> tuple[str, list[ExtractedVisual]]:
    """Replace base64 image nodes with compact handles using markdown-it tokens."""
    parser = MarkdownIt("zero").enable("image")
    visuals: list[ExtractedVisual] = []
    handles_by_uri: dict[str, str] = {}
    for token in parser.parse(text):
        if token.type != "inline" or not token.children:
            continue
        for child in token.children:
            if child.type != "image":
                continue
            src = child.attrGet("src")
            if not isinstance(src, str) or not src.startswith("data:"):
                continue
            decoded = _decode_data_uri(src)
            if decoded is None:
                continue
            media_type, data = decoded
            handle_id = handles_by_uri.get(src)
            if handle_id is None:
                handle_id = _mint_handle_id(data)
                handles_by_uri[src] = handle_id
                visuals.append(
                    ExtractedVisual(
                        handle_id=handle_id,
                        anchor=child.content or None,
                        media_type=media_type,
                        data=data,
                    )
                )
            text = text.replace(src, f"visual://{handle_id}")
    return text, visuals


def _decode_data_uri(uri: str) -> tuple[str, bytes] | None:
    header, _, payload = uri[len("data:") :].partition(",")
    if not payload:
        return None
    parameters = header.split(";")
    if "base64" not in parameters[1:]:
        return None
    media_type = parameters[0] or "application/octet-stream"
    try:
        data = base64.b64decode(payload, validate=True)
    except binascii.Error, ValueError:
        return None
    return media_type, data


def _extract_xlsx_visuals(content: bytes) -> list[ExtractedVisual]:
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    visuals: list[ExtractedVisual] = []
    for worksheet in workbook.worksheets:
        for image in getattr(worksheet, "_images", []):
            data = image._data()
            visuals.append(
                ExtractedVisual(
                    handle_id=_mint_handle_id(data),
                    anchor=_xlsx_anchor(worksheet.title, image),
                    media_type=f"image/{(image.format or 'png').lower()}",
                    data=data,
                )
            )
    return visuals


def _xlsx_anchor(sheet_title: str, image: object) -> str | None:
    marker = getattr(getattr(image, "anchor", None), "_from", None)
    if marker is None:
        return None
    cell = f"{get_column_letter(marker.col + 1)}{marker.row + 1}"
    return f"{sheet_title}!{cell}"


def _mint_handle_id(data: bytes) -> str:
    return f"vis-{sha256(data).hexdigest()[:16]}"


__all__ = [
    "ConvertedResource",
    "ExtractedVisual",
    "ResourceConversionError",
    "UnsafeArchiveError",
    "convert_resource",
    "is_convertible",
]
