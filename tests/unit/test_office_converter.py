# Copyright 2025-2026 Hanlian Lu. SPDX-License-Identifier: Apache-2.0
"""Tests for LibreOffice converter."""

from __future__ import annotations

import inspect
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from dlightrag.config import DlightragConfig
from dlightrag.converters.office import LibreOfficeConverter, OfficeConverterError, PageSetup


class TestLibreOfficeConverter:
    """Test converter configuration and should_convert logic."""

    def _make_converter(self, test_config: DlightragConfig) -> LibreOfficeConverter:
        return LibreOfficeConverter(test_config)

    def test_should_convert_xlsx(self, test_config: DlightragConfig) -> None:
        """Test that .xlsx triggers conversion."""
        converter = self._make_converter(test_config)
        assert converter.should_convert(Path("test.xlsx"))
        assert converter.should_convert(Path("test.xls"))

    def test_should_not_convert_pdf(self, test_config: DlightragConfig) -> None:
        """Test that .pdf does not trigger conversion."""
        converter = self._make_converter(test_config)
        assert not converter.should_convert(Path("test.pdf"))
        assert not converter.should_convert(Path("test.docx"))

    def test_should_skip_csv(self, test_config: DlightragConfig) -> None:
        """Test that .csv is skipped."""
        converter = self._make_converter(test_config)
        assert not converter.should_convert(Path("test.csv"))

    def test_respects_config_flag(self, tmp_path: Path) -> None:
        """Test that excel_auto_convert_to_pdf=False disables conversion."""
        config = DlightragConfig(  # type: ignore[call-arg]
            working_dir=str(tmp_path),
            openai_api_key="test-key",
            excel_auto_convert_to_pdf=False,
            kv_storage="JsonKVStorage",
            doc_status_storage="JsonDocStatusStorage",
            vector_storage="NanoVectorDBStorage",
            graph_storage="NetworkXStorage",
        )
        converter = LibreOfficeConverter(config)
        assert not converter.should_convert(Path("test.xlsx"))

    def test_docling_parser_skips_conversion(self, tmp_path: Path) -> None:
        """Test that docling parser disables Excel conversion."""
        config = DlightragConfig(  # type: ignore[call-arg]
            working_dir=str(tmp_path),
            openai_api_key="test-key",
            parser="docling",
            kv_storage="JsonKVStorage",
            doc_status_storage="JsonDocStatusStorage",
            vector_storage="NanoVectorDBStorage",
            graph_storage="NetworkXStorage",
        )
        converter = LibreOfficeConverter(config)
        assert not converter.should_convert(Path("test.xlsx"))

    def test_is_safe_to_delete(self, test_config: DlightragConfig) -> None:
        """Test safety check for file deletion."""
        converter = self._make_converter(test_config)

        # File within storage is safe
        safe_path = test_config.working_dir_path / "sources" / "test.xlsx"
        assert converter._is_safe_to_delete(safe_path)

        # File outside storage is not safe
        assert not converter._is_safe_to_delete(Path("/tmp/outside.xlsx"))


class TestEstimateExcelPageWidth:
    """Test adaptive page width estimation from Excel column widths."""

    def _make_converter(self, test_config):
        return LibreOfficeConverter(test_config)

    def _make_xlsx(self, tmp_path, columns, sheet_configs=None):
        """Helper: create xlsx with specified column widths."""
        wb = openpyxl.Workbook()
        if sheet_configs:
            for i, cfg in enumerate(sheet_configs):
                ws = wb.active if i == 0 else wb.create_sheet(cfg["name"])
                if i == 0:
                    ws.title = cfg["name"]
                for col_letter, width in cfg["columns"]:
                    ws.column_dimensions[col_letter].width = width
                    ws.cell(
                        row=1,
                        column=openpyxl.utils.column_index_from_string(col_letter),
                        value="data",
                    )
        else:
            ws = wb.active
            for col_letter, width in columns:
                ws.column_dimensions[col_letter].width = width
                ws.cell(
                    row=1,
                    column=openpyxl.utils.column_index_from_string(col_letter),
                    value="data",
                )
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        wb.close()
        return path

    def test_narrow_excel_gives_portrait(self, test_config, tmp_path):
        """3 narrow columns -> total ~6cm + margins -> clamped to 21cm -> portrait."""
        path = self._make_xlsx(tmp_path, [("A", 10), ("B", 10), ("C", 10)])
        converter = self._make_converter(test_config)
        setup = converter._estimate_excel_page_width(path)
        assert isinstance(setup, PageSetup)
        assert setup.orientation == "portrait"
        assert setup.width_cm == 21.0
        assert setup.height_cm == 29.7

    def test_wide_excel_gives_landscape(self, test_config, tmp_path):
        """20 columns at width 15 -> ~60cm + margins -> landscape."""
        cols = [(openpyxl.utils.get_column_letter(i + 1), 15) for i in range(20)]
        path = self._make_xlsx(tmp_path, cols)
        converter = self._make_converter(test_config)
        setup = converter._estimate_excel_page_width(path)
        assert setup.orientation == "landscape"
        assert setup.width_cm > 21.0
        assert setup.width_cm <= 63.0
        assert setup.height_cm == 21.0

    def test_very_wide_clamped_to_max(self, test_config, tmp_path):
        """50 columns at width 20 -> way over 63cm -> clamped to 63."""
        cols = [(openpyxl.utils.get_column_letter(i + 1), 20) for i in range(50)]
        path = self._make_xlsx(tmp_path, cols)
        converter = self._make_converter(test_config)
        setup = converter._estimate_excel_page_width(path)
        assert setup.width_cm == 63.0

    def test_hidden_columns_excluded(self, test_config, tmp_path):
        """Hidden columns should not contribute to width."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 100
        ws.column_dimensions["B"].hidden = True
        ws.column_dimensions["C"].width = 10
        ws.cell(row=1, column=1, value="a")
        ws.cell(row=1, column=2, value="b")
        ws.cell(row=1, column=3, value="c")
        path = tmp_path / "hidden.xlsx"
        wb.save(str(path))
        wb.close()

        converter = self._make_converter(test_config)
        setup = converter._estimate_excel_page_width(path)
        assert setup.width_cm == 21.0
        assert setup.orientation == "portrait"

    def test_multi_sheet_uses_widest(self, test_config, tmp_path):
        """Width estimation should use the widest sheet."""
        path = self._make_xlsx(
            tmp_path,
            [],
            sheet_configs=[
                {"name": "Narrow", "columns": [("A", 10), ("B", 10)]},
                {
                    "name": "Wide",
                    "columns": [
                        (openpyxl.utils.get_column_letter(i + 1), 15) for i in range(20)
                    ],
                },
            ],
        )
        converter = self._make_converter(test_config)
        setup = converter._estimate_excel_page_width(path)
        assert setup.width_cm > 21.0
        assert setup.orientation == "landscape"

    def test_xls_fallback(self, test_config, tmp_path):
        """xls files fall back to A4 landscape."""
        path = tmp_path / "old.xls"
        path.write_bytes(b"fake xls content")
        converter = self._make_converter(test_config)
        setup = converter._estimate_excel_page_width(path)
        assert setup.width_cm == 29.7
        assert setup.height_cm == 21.0
        assert setup.orientation == "landscape"

    def test_corrupt_file_fallback(self, test_config, tmp_path):
        """Corrupt xlsx falls back to A4 landscape."""
        path = tmp_path / "corrupt.xlsx"
        path.write_bytes(b"not a real xlsx")
        converter = self._make_converter(test_config)
        setup = converter._estimate_excel_page_width(path)
        assert setup.width_cm == 29.7
        assert setup.height_cm == 21.0
        assert setup.orientation == "landscape"


class TestFileToPdfBytes:
    """Test file_to_pdf_bytes method."""

    def _make_converter(self, test_config):
        return LibreOfficeConverter(test_config)

    def test_excel_delegates_to_convert_excel_to_pdf(self, test_config, tmp_path):
        """xlsx files should go through the Excel->ODS->PDF pipeline."""
        converter = self._make_converter(test_config)
        xlsx_path = tmp_path / "test.xlsx"
        xlsx_path.touch()

        fake_pdf = tmp_path / "test.pdf"
        fake_pdf.write_bytes(b"%PDF-1.4 fake")

        with patch.object(converter, "_convert_excel_to_pdf", return_value=fake_pdf):
            result = converter.file_to_pdf_bytes(xlsx_path)

        assert result == b"%PDF-1.4 fake"

    def test_docx_uses_standard_libreoffice(self, test_config, tmp_path):
        """docx files should use standard libreoffice conversion."""
        converter = self._make_converter(test_config)
        docx_path = tmp_path / "doc.docx"
        docx_path.touch()

        mock_result = MagicMock()
        mock_result.returncode = 0

        def fake_run(cmd, **kwargs):
            outdir = cmd[cmd.index("--outdir") + 1]
            (Path(outdir) / "doc.pdf").write_bytes(b"%PDF-docx")
            return mock_result

        with patch("dlightrag.converters.office.subprocess.run", side_effect=fake_run):
            result = converter.file_to_pdf_bytes(docx_path)

        assert result == b"%PDF-docx"

    def test_nonexistent_file_raises(self, test_config, tmp_path):
        """Missing file should raise OfficeConverterError."""
        converter = self._make_converter(test_config)
        with pytest.raises(OfficeConverterError, match="Source file not found"):
            converter.file_to_pdf_bytes(tmp_path / "missing.xlsx")

    def test_failed_conversion_raises(self, test_config, tmp_path):
        """Failed LibreOffice conversion should raise OfficeConverterError."""
        converter = self._make_converter(test_config)
        pptx_path = tmp_path / "slides.pptx"
        pptx_path.touch()

        mock_result = MagicMock()
        mock_result.returncode = 1
        mock_result.stderr = "error"

        with patch("dlightrag.converters.office.subprocess.run", return_value=mock_result):
            with pytest.raises(OfficeConverterError, match="conversion failed"):
                converter.file_to_pdf_bytes(pptx_path)


class TestConvertBytesToPdfAdaptive:
    """Test that convert_bytes_to_pdf uses adaptive width for Excel."""

    def _make_converter(self, test_config):
        return LibreOfficeConverter(test_config)

    def test_excel_bytes_uses_adaptive_pipeline(self, test_config, tmp_path):
        """Excel MIME type should trigger adaptive conversion."""
        converter = self._make_converter(test_config)

        fake_pdf_path = tmp_path / "output.pdf"
        fake_pdf_path.write_bytes(b"%PDF-adaptive")

        with patch.object(converter, "_convert_excel_to_pdf", return_value=fake_pdf_path) as mock:
            result = converter.convert_bytes_to_pdf(
                b"fake excel bytes",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        assert result == b"%PDF-adaptive"
        mock.assert_called_once()

    def test_apply_page_setup_param_removed(self, test_config):
        """apply_page_setup parameter should no longer exist."""
        converter = self._make_converter(test_config)
        sig = inspect.signature(converter.convert_bytes_to_pdf)
        assert "apply_page_setup" not in sig.parameters


class TestSetOdsPageSetup:
    """Test ODS XML page setup modification."""

    def _make_converter(self, test_config):
        return LibreOfficeConverter(test_config)

    def _make_minimal_ods(self, tmp_path):
        """Create a minimal ODS file with styles.xml for testing."""
        ods_path = tmp_path / "test.ods"
        extract_dir = tmp_path / "ods_build"
        extract_dir.mkdir()

        (extract_dir / "mimetype").write_text("application/vnd.oasis.opendocument.spreadsheet")

        styles_xml = extract_dir / "styles.xml"
        styles_xml.write_text(
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<office:document-styles '
            'xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0" '
            'xmlns:style="urn:oasis:names:tc:opendocument:xmlns:style:1.0" '
            'xmlns:fo="urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0">'
            '<office:automatic-styles>'
            '<style:page-layout style:name="pm1">'
            '<style:page-layout-properties fo:page-width="21cm" fo:page-height="29.7cm" '
            'style:print-orientation="portrait"/>'
            '</style:page-layout>'
            '</office:automatic-styles>'
            '</office:document-styles>'
        )

        with zipfile.ZipFile(ods_path, "w") as zf:
            zf.write(extract_dir / "mimetype", "mimetype", compress_type=zipfile.ZIP_STORED)
            zf.write(styles_xml, "styles.xml", compress_type=zipfile.ZIP_DEFLATED)

        return ods_path

    def test_sets_custom_dimensions(self, test_config, tmp_path):
        """Verify page width, height, orientation, margins are written to ODS."""
        ods_path = self._make_minimal_ods(tmp_path)
        converter = self._make_converter(test_config)
        setup = PageSetup(width_cm=45.0, height_cm=21.0, orientation="landscape")

        converter._set_ods_page_setup(ods_path, setup)

        extract_dir = tmp_path / "verify"
        with zipfile.ZipFile(ods_path, "r") as zf:
            zf.extractall(extract_dir)

        tree = ET.parse(extract_dir / "styles.xml")
        ns = {
            "style": "urn:oasis:names:tc:opendocument:xmlns:style:1.0",
            "fo": "urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0",
        }
        props = tree.getroot().find(".//style:page-layout-properties", ns)
        assert props is not None

        fo = "urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0"
        style = "urn:oasis:names:tc:opendocument:xmlns:style:1.0"

        assert props.get(f"{{{fo}}}page-width") == "45.0cm"
        assert props.get(f"{{{fo}}}page-height") == "21.0cm"
        assert props.get(f"{{{style}}}print-orientation") == "landscape"
        assert props.get(f"{{{fo}}}margin-left") == "0.5cm"
        assert props.get(f"{{{fo}}}margin-right") == "0.5cm"
        assert props.get(f"{{{fo}}}margin-top") == "0.5cm"
        assert props.get(f"{{{fo}}}margin-bottom") == "0.5cm"
        assert props.get(f"{{{style}}}scale-to-X") == "1"
        assert props.get(f"{{{style}}}scale-to-Y") == "0"

    def test_portrait_setup(self, test_config, tmp_path):
        """Verify portrait orientation is set correctly."""
        ods_path = self._make_minimal_ods(tmp_path)
        converter = self._make_converter(test_config)
        setup = PageSetup(width_cm=21.0, height_cm=29.7, orientation="portrait")

        converter._set_ods_page_setup(ods_path, setup)

        extract_dir = tmp_path / "verify"
        with zipfile.ZipFile(ods_path, "r") as zf:
            zf.extractall(extract_dir)

        tree = ET.parse(extract_dir / "styles.xml")
        ns = {"style": "urn:oasis:names:tc:opendocument:xmlns:style:1.0"}
        props = tree.getroot().find(".//style:page-layout-properties", ns)
        assert props is not None
        assert props.get(f"{{{ns['style']}}}print-orientation") == "portrait"
